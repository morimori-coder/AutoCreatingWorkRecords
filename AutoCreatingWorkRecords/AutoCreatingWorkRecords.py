from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import GetJson
import time
import const
import bs4 
import pyautogui
import datetime
import re
import openpyxl

loginInfo = GetJson.get_json_data(const.LOGIN_INFO)

# chrome driverでGoogle Chromeを開く
driver = webdriver.Chrome(executable_path = loginInfo["chromeDriver"])

# ログインしたいページのURLを入力する
driver.get(loginInfo["url"])
time.sleep(3)

# ID,パスワードを入力する
loginID = driver.find_element_by_id("username")
loginID.send_keys(loginInfo["id"])
loginPass = driver.find_element_by_id("password")
loginPass.send_keys(loginInfo["password"])

# seleniumで対象のページへログインする
loginButton = driver.find_element_by_id('Login')
loginButton.click()

elementsNames = GetJson.get_json_data(const.ELEMENTS_NAME)

# 勤務表を開く
workRecordTab = driver.find_element_by_id(elementsNames["workedRecord"])
workRecordTab.click()
time.sleep(2)

# 月次サマリーをクリック
monthSummery = driver.find_element_by_class_name(elementsNames['monthlySum'])
monthSummery.click()
time.sleep(2)

pyautogui.hotkey('ctrl','s')
time.sleep(2)
fine_name = datetime.datetime.now().strftime('%Y%m%d') + 'wordRecord.html'
pyautogui.typewrite(fine_name)
pyautogui.press("enter")
time.sleep(3)

#ローカルに保存した勤務表を開く
fine_name = datetime.datetime.now().strftime('%Y%m%d') + 'wordRecord.html'
savePath = GetJson.get_json_data(const.SAVE_PATH)
soup = bs4.BeautifulSoup(open(savePath["folderPath"] + fine_name, encoding = 'utf-8'), 'html.parser')
#勤務表の要素を取得する
element = soup.find_all('tr', attrs = {'class' : re.compile('prtv.*')})

#Excelの勤務表を開く
wb = openpyxl.load_workbook(savePath["excelPath"])
ws = wb["勤怠管理表"]

dateInfo = datetime.datetime.now()
#B2にyyyy年mm月
#ws["B2"].value = str(dateInfo.year) + "年" + str(dateInfo.month) + "月"
ws["B2"].value = dateInfo.strftime('%Y/%m/') + '1'
#C5にyyyy年
ws["C5"].value = str(dateInfo.year) + "年"
#C6にmm月度
ws["C6"].value = str(dateInfo.month) + "月度"

rowNum = 14
for row in element:
    if rowNum > 44 :
        break
    #C14～には有給(全休)、有給(阪急)、欠勤、遅刻、早退のいずれか
    ws["C" + str(rowNum)].value = ''
    ws["D" + str(rowNum)].value = ''
    ws["E" + str(rowNum)].value = ''
    ws["F" + str(rowNum)].value = ''
    ws["N" + str(rowNum)].value = ''

    if row.contents[5].string is None:
        rowNum += 1
        continue
    #Dは始業時間hh:mm
    tdatetime = datetime.datetime.strptime(row.contents[5].string + ':00', '%H:%M:%S')
    ws["D" + str(rowNum)].value = tdatetime.time()

    #Eは退勤時間hh:mm
    tdatetime = datetime.datetime.strptime(row.contents[6].string + ':00', '%H:%M:%S')
    ws["E" + str(rowNum)].value = tdatetime.time()

    #Fは休憩h:mm
    tdatetime = datetime.datetime.strptime(row.contents[9].string + ':00', '%H:%M:%S')
    ws["F" + str(rowNum)].value = tdatetime.time()

    rowNum += 1

driver.quit()
wb.save(savePath["excelPath"])